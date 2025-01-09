from openpyxl import load_workbook
from PIL import Image

wb = load_workbook('./raw0.xlsx')
ws = wb[wb.sheetnames[0]]
# collect pictures and rename using its row
"""
for image in ws._images:
    #remember to change the path to the repository assets folder
    img_name='./assets/person_img/row'+str(image.anchor._from.row-1)+'.png'
    img=Image.open(image.ref).convert('RGB')
    img.save(img_name,"PNG")"""

#row_flag
flag_row=ws[2]
for col in flag_row:
    if col.value=='姓名':
        name_cn_flag=col.col_idx-1
    elif col.value=='英文姓名':
        name_en_flag=col.col_idx-1
    elif col.value=='slogan中文':
        slogan_cn_flag=col.col_idx-1
    elif col.value=='slogan英文':
        slogan_en_flag=col.col_idx-1
    elif col.value=='想跳转的个人链接（可填无）':
        link_flag=col.col_idx-1

# calculate the mode of the image size
with open('./assets/person_img/mode.txt','w') as fp:
    for i in range(3,ws.max_row):
        img=Image.open('./assets/person_img/row'+str(i-2)+'.png')
        w=img.size[0]
        h=img.size[1]
        quot=h/w
        if quot>=1.0:
            fp.write('1\n')
        else:
            fp.write('2\n')

# collect data and write to markdown
row_index=1
quot_txt=open('./assets/person_img/mode.txt','r')
for row in ws.iter_rows(min_row=row_index+2, max_row=ws.max_row-1,values_only=True):
    filename='./_projects/'+str(row_index)+'.md'  #remember to change the path to the repository _projecs folder
    imgpath='assets/person_img/row'+str(row_index)+'.png'
    with open(filename, 'w', encoding='utf-8') as fp:
        fp.write('---\n')
        fp.write('layout: page\n')
        fp.write('title: '+str(row[name_cn_flag])+'\n')
        print(str(row[name_cn_flag]))
        fp.write('description: '+str(row[name_en_flag])+'\n')
        print(str(row[name_en_flag]))
        fp.write('img: '+imgpath+'\n')
        fp.write('importance: '+str(row_index)+'\n')
        fp.write('category: student\n')
        fp.write('---\n\n')
        mode_f=quot_txt.readline()
        if mode_f=='1\n':
            fp.write('<div class="row justify-content-center">\n')
            fp.write('    <div class="col-4 mt-3 mt-md-0">\n')
            fp.write('        {% include figure.liquid loading="eager" path="'+imgpath+'" title="example image" class="img-fluid rounded z-depth-1" %}\n')
            fp.write('    </div>\n')
            fp.write('</div>\n\n')
        else:
            fp.write('<div class="row justify-content-center">\n')
            fp.write('    <div class="col-6 mt-3 mt-md-0">\n')
            fp.write('        {% include figure.liquid loading="eager" path="'+imgpath+'" title="example image" class="img-fluid rounded z-depth-1" %}\n')
            fp.write('    </div>\n')
            fp.write('</div>\n\n')
        fp.write('<font size="5">\n')
        fp.write('    name: '+str(row[name_cn_flag])+'<br>\n')
        fp.write('    name_en: '+str(row[name_en_flag])+'<br>\n')
        if row[slogan_cn_flag] != '无':
            fp.write('    slogan: '+str(row[slogan_cn_flag])+'<br>\n')
        if row[slogan_en_flag] != '无':
            fp.write('    slogan_en: '+str(row[slogan_en_flag])+'<br>\n')
        if row[link_flag] != '无':
            fp.write('    link: <a href="'+str(row[link_flag])+'">'+str(row[link_flag])+'</a><br>\n')
        fp.write('</font>\n')
    row_index+=1
quot_txt.close()